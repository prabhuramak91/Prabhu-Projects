
# coding: utf-8

# In[1]:

#importing packages
import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
import csv
import numpy as np


# In[2]:

#Link for the medicare data
url = "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"


# In[3]:

#Downloading the file
r = requests.get(url)


# In[4]:

#Creating a staging directory
staging_dirname = "staging"
os.mkdir(staging_dirname)


# In[5]:

#Copying contents of r to staging directory
zipfilename = os.path.join(staging_dirname, "test.zip")
z = open(zipfilename, 'wb')
z.write(r.content)
z.close()


# In[6]:

#Extracting the zip files to staging directory
x = zipfile.ZipFile(zipfilename, "r")
x.extractall(staging_dirname)
x.close()


# In[7]:

#Making a list of csv files
Files = []
Names = []
Direct = []
glob_dirname =  os.path.join(staging_dirname, "*.csv")
#Creating Files and path of the csv files
for i in glob.glob(glob_dirname):
    g = os.path.basename(i)
    h = os.path.splitext(os.path.basename(i))[0]
    Direct.append(i)
    Files.append(g)
    Names.append(h)   
alpha = list('abcdefghijklmnopqrstuvwxyz')
#Removing the corrupted file from the list
Files.remove("FY2015_Percent_Change_in_Medicare_Payments.csv")
Names.remove("FY2015_Percent_Change_in_Medicare_Payments")
Direct.remove(os.path.join(staging_dirname,"FY2015_Percent_Change_in_Medicare_Payments.csv"))


# In[8]:

#Loop for encoding all files
for File in Files:
    #Reading cp1252 encoded files 
    fn = os.path.join(staging_dirname, File )
    in_fp = open(fn, 'rt', encoding = 'cp1252')
    inp = in_fp.read()
    in_fp.close()
    #writing them out in utf-8
    ofn = os.path.join(staging_dirname, File)
    out_fp = open(ofn, 'wt', encoding = 'utf-8')
    for c in inp:
        if c != '\0':
            out_fp.write(c)
    out_fp.close()
    


# In[9]:

#Converting the table name into required format
def tablename(Name):
    """
    Changing the csv filename into tablename
    """
    
    Name = Name.lower()
    Name = Name.replace(" ","_")
    Name = Name.replace("-","_")
    Name = Name.replace("/","_")
    Name = Name.replace("%","pct")
    #Special case for names with non alphabetic start
    if Name[0] not in alpha:
        Name = "t_" + Name
    return Name    
    


# In[10]:

#Converting the column name into required format
def colname(Name):
    """
    Changing the csvfile columnnames into db columnnnames
    """
    Name = Name.lower()
    Name = Name.replace(" ","_")
    Name = Name.replace("-","_")
    Name = Name.replace("/","_")
    Name = Name.replace("%","pct")
    #Special case for names with non alphabetic start
    if Name[0] not in alpha:
        Name = "c_" + Name
    return Name    
    


# In[17]:

#Creating a database through sqlite3
conn = sqlite3.connect("medicare_hospital_compare.db")
c1 = conn.cursor()
c2 = conn.cursor()
#Loop for each csv file to be converted into a table
for i in range(0,len(Direct)):
    Colname=[]
    f = open(Direct[i], encoding = 'utf-8')
    reader = csv.reader(f)
    col = next(reader)
    Tblname = tablename(Names[i])
    #Creating a table for each csv file in the database
    for Col in col:
        Colname.append(colname(Col))   
    sql = """create table if not exists """+Tblname+" ("+",".join([("%s text" %name) for name in Colname])+ ")"
    c1.execute(sql)
    gist = []
    #Inserting values into the tables created
    #Removing last two lines for corrupted files
    if Files[i] in ["MORT_READM_April2017.csv","PSI_April2017.csv"]:
        for line in reader:
            gist.append(line)
        gist = gist[:-1]
        gist = gist[:-1]    
        for line in gist:
            sql_str = "insert into "+Tblname+ " ("+",".join([("%s " %name) for name in Colname])+") values ("+",".join(["?" for i in range(0, len(Colname))])+")"
            sql_tuple = tuple(line)
            c1.execute(sql_str, sql_tuple)
    #Normal insert operation for other files        
    else:
        for line in reader:
            gist.append(line)
        for line in gist:
            sql_str = "insert into "+Tblname+ " ("+",".join([("%s " %name) for name in Colname])+") values ("+",".join(["?" for i in range(0, len(Colname))])+")"
            sql_tuple = tuple(line)
            c1.execute(sql_str, sql_tuple)                       


# In[18]:

#Link for downloading hospital ranking workbook
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"


# In[19]:

#Downloading the file 
p = requests.get(k_url)
#Writing into a excel workbook
y = open("hospital_ranking_focus_states.xlsx", 'wb')
y.write(p.content)
y.close()


# In[20]:

#Writing the xlsx file into a database table
mx = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
sheet = mx.get_sheet_by_name("Hospital National Ranking")
#Extracting Column names and values to be inserted
columns = [sheet.cell(row=1, column=1).value,sheet.cell(row=1,column=2).value]
i = 2
sqlist = []
#Writing the value into each cell
while sheet.cell(row=i, column=1).value != None:
    sqlist.append([sheet.cell(row=i, column=1).value,sheet.cell(row=i,column=2).value])
    i+=1


# In[21]:

#Getting the no of columns for creating a table
#Creating a table for Hospital national ranking worksheet in database    
sql_st = "create table if not exists "+tablename("Hospital National Ranking")+" (provider_id text, ranking number)"
c2.execute(sql_st)
for sqist in sqlist: 
    sql_s = "insert into "+tablename("Hospital National Ranking")+ " (provider_id, ranking)" + " values ("'?'", "'?'")"
    sql_tup = tuple(sqist) 
    c2.execute(sql_s,sql_tup)
#Closing the database connection    
conn.commit()
conn.close()


# In[22]:

#Creating a workbook hospital_ranking.xlsx
con = sqlite3.connect("medicare_hospital_compare.db")
c3 = con.cursor()
c4 = con.cursor()
wb1 = openpyxl.Workbook()
wb1.remove_sheet(wb1.get_sheet_by_name('Sheet'))
#Column names for the workbook
lost = ['Provider ID', 'Hospital Name', 'City', 'State', 'County']
#Sql query for each worksheet
sql1 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id order by ranking limit 100"
sql2 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'CA' order by ranking limit 100"
sql3 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'FL' order by ranking limit 100"
sql4 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'GA' order by ranking limit 100"
sql5 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'IL' order by ranking limit 100"
sql6 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'KS' order by ranking limit 100"
sql7 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'MI' order by ranking limit 100"
sql8 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'NY' order by ranking limit 100"
sql9 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'OH' order by ranking limit 100"
sql10 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'PA' order by ranking limit 100"
sql11 = "select distinct n.provider_id, hospital_name, city, state, county_name from hospital_general_information g, hospital_national_ranking n where g.provider_id = n.provider_id and state = 'TX' order by ranking limit 100"
#Creating a dictionary of worksheet names

dit = {sql1:'Nationwide', sql2:'California', sql3:'Florida', sql4:'Georgia', sql5:'Illinois', sql6:'Kansas', sql7:'Michigan', sql8:'New York', sql9:'Ohio', sql10:'Pennsylvania', sql11:'Texas'}
#Exporting the data for each worksheet
for sql in dit.keys():
    ro = ()
    Nation = c3.execute(sql)
    for row in Nation:
        ro = ro + row
    sheet1 = wb1.create_sheet(dit[sql])    
    for i in range(5):
        sheet1.cell(row = 1, column = i+1, value = lost[i])
    for i in range(1,101):
        for j in range(1,6):
            sheet1.cell(row = i+1, column = j, value = ro[(i-1)*5+(j-1)])
#Saving the worksheet with desired name and closing it            
wb1.save("hospital_ranking.xlsx")
wb1.close()


# In[23]:

#Creating a VIEW
s1 = "select measure_id, measure_name from timely_and_effective_care___hospital group by measure_id, measure_name"
measurelist = c4.execute(s1)
mid = []
mname = []
for row in measurelist:
    mid.append(row[0])
    mname.append(row[1])
del mid[0]
del mname[0]
#Dictionary of measure_id with measure_name
dic = {mid[i]:mname[i] for i in range(21)}
s2 = "create view if not exists measure as select measure_id, measure_name, cast(score as int) as sco, state from timely_and_effective_care___hospital where score != 'Not Available' and length(score)<7"
c4.execute(s2)


# In[24]:

#Function to return no of measure_id
def med(q):
    """
    Returns no of distinct measure_id
    """
    mod = []
    bod = c4.execute(q)
    for md in bod:
        mod = mod + list(md)
    return(mod)


# In[25]:

#Creating a Measure_Statistics.xlsx
wb2 = openpyxl.Workbook()
wb2.remove_sheet(wb2.get_sheet_by_name('Sheet'))
#Creating a list for column names of the worksheet
post = ['Measure ID', 'Measure Name', 'Minimum', 'Maximum', 'Average', 'Standard Deviation']
#Queries for each focus states
sq1 = "select sco from measure where "
sq2 = "select sco from measure where state = 'CA' and "
sq3 = "select sco from measure where state = 'FL' and "
sq4 = "select sco from measure where state = 'GA' and "
sq5 = "select sco from measure where state = 'IL' and "
sq6 = "select sco from measure where state = 'KS' and "
sq7 = "select sco from measure where state = 'MI' and "
sq8 = "select sco from measure where state = 'NY' and "
sq9 = "select sco from measure where state = 'OH' and "
sq10 = "select sco from measure where state = 'PA' and "
sq11 = "select sco from measure where state = 'TX' and "
#Using med function to get distinct measure_id for each focus state
q1 = mid
q2 = med("select distinct measure_id from measure where state = 'CA' order by measure_id")
q3 = med("select distinct measure_id from measure where state = 'FL' order by measure_id")
q4 = med("select distinct measure_id from measure where state = 'GA' order by measure_id")
q5 = med("select distinct measure_id from measure where state = 'IL' order by measure_id")
q6 = med("select distinct measure_id from measure where state = 'KS' order by measure_id")
q7 = med("select distinct measure_id from measure where state = 'MI' order by measure_id")
q8 = med("select distinct measure_id from measure where state = 'NY' order by measure_id")
q9 = med("select distinct measure_id from measure where state = 'OH' order by measure_id")
q10 = med("select distinct measure_id from measure where state = 'PA' order by measure_id")
q11 = med("select distinct measure_id from measure where state = 'TX' order by measure_id")
#Creating dictionaries of query and sheet name
de = {sq1:q1, sq2:q2, sq3:q3, sq4:q4, sq5:q5, sq6:q6, sq7:q7, sq8:q8, sq9:q9, sq10:q10, sq11:q11}
di = {sq1:'Nationwide', sq2:'California', sq3:'Florida', sq4:'Georgia', sq5:'Illinois', sq6:'Kansas', sq7:'Michigan', sq8:'New York', sq9:'Ohio', sq10:'Pennsylvania', sq11:'Texas'}
#Querying the values from the view table
for sl in di.keys():
    po = []
    for i in de[sl]:
        sql = sl + "measure_id = '%s'" %i 
        nat = c4.execute(sql)
        score = []
        for n in nat:
            score = score + list(n)
        sc = np.asarray(score)
        po = po + [i, dic[i], np.min(sc), np.max(sc), np.mean(sc), np.std(sc)]
    sheet1 = wb2.create_sheet(di[sl])
    #Inserting values into each cell in the worksheet
    for j in range(6):
        sheet1.cell(row = 1, column = j+1, value = post[j])
    for j in range(1, len(de[sl])+1):
        for k in range(1,7):
            sheet1.cell(row = j+1, column = k, value = po[(j-1)*6+(k-1)])   
#saving the worksheet            
wb2.save('measures_statistics.xlsx')
wb2.close()


# In[26]:

#Closing the database
con.commit()
con.close()

